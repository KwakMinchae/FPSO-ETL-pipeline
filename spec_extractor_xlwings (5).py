"""
Spec Sheet Data Extractor — xlwings version
============================================
Reads directly from Excel while it is open.
No need to remove encryption or make copies.

HOW TO USE:
  1. Install xlwings once:
         py -m pip install xlwings

  2. Open your .xlsm file in Excel (enter the password if prompted)

  3. Run this script in Command Prompt:
         py spec_extractor_xlwings.py

  4. Type the exact filename when prompted (e.g. project_file.xlsm)

  5. Output file appears in the same folder as this script.
"""

from __future__ import annotations
import os
import re
import sys
from typing import Sequence, Tuple

try:
    import xlwings as xw
except ImportError:
    print("xlwings is not installed. Run this first:")
    print("    py -m pip install xlwings")
    input("\nPress Enter to exit...")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl is not installed. Run this first:")
    print("    py -m pip install openpyxl")
    input("\nPress Enter to exit...")
    sys.exit(1)

# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

SKIP_SHEETS = {"ds"}
OUTPUT_SHEET_NAME = "ds"

FIELDS: Sequence[Tuple[str, Tuple[int, int]]] = [
    # ── GENERAL ──────────────────────────────────────────────────────────────
    ("Tag Number",              (1,  21)),
    ("PID",                     (2,  21)),
    ("Service",                 (3,  21)),
    ("Line Number",             (4,  21)),
    ("Line Size",               (4,  30)),
    ("Insulation",              (4,  32)),
    ("Pipe Class",              (4,  36)),
    ("SIL Rating",              (5,  21)),
    ("Ex Cert",                 (5,  29)),
    ("ATEX Marking",            (5,  33)),

    # ── PROCESS CONDITIONS ───────────────────────────────────────────────────
    ("Fluid",                   (6,  21)),
    ("Phase",                   (6,  29)),
    ("Pressure Unit",           (7,  21)),
    ("Design/Shutoff Pressure", (7,  23)),
    ("Temp Unit",               (8,  21)),
    ("Design Temp Min",         (8,  23)),
    ("Design Temp Max",         (8,  30)),
    ("Op Pressure Normal",      (9,  23)),
    ("Op Pressure Max",         (9,  32)),
    ("Op Temp Min",             (10, 23)),
    ("Op Temp Normal",          (10, 28)),
    ("Op Temp Max",             (10, 33)),
    ("Methanol Content",        (11, 23)),
    ("NACE",                    (12, 14)),
    ("Solid Particles",         (12, 21)),

    # ── BALL VALVE / PIPING ──────────────────────────────────────────────────
    ("Service Class",           (14, 21)),
    ("Standard",                (14, 24)),
    ("Leakage",                 (14, 31)),
    ("Bore Type",               (15, 21)),
    ("ID",                      (15, 31)),
    ("Trim Design",             (16, 21)),
    ("Metal to Metal Seat",     (16, 29)),
    ("Stem Design",             (17, 21)),
    ("Body Construction",       (17, 29)),
    ("Body Material",           (18, 21)),
    ("Ball Material",           (18, 29)),
    ("Stem Material",           (19, 21)),
    ("Seat Material",           (19, 29)),
    ("Dynamic Seal Material",   (20, 21)),
    ("Static Seal Material",    (20, 29)),
    ("Insert Material",         (21, 21)),
    ("Back Up Material",        (21, 29)),
    ("Sealant Injection",       (22, 21)),
    ("Bypass",                  (22, 27)),
    ("Stuff Box Packing",       (22, 33)),
    ("Drain/Vents",             (23, 21)),
    ("Anti Static",             (23, 29)),
    ("Locking Device",          (24, 21)),
    ("Bolting Material",        (24, 29)),
    ("Supports",                (25, 21)),
    ("Lifting Lug",             (25, 29)),
    ("Corrosion Allowance",     (26, 21)),
    ("Weld Overlay",            (26, 29)),
    ("Insulation (BV)",         (27, 21)),  # e.g. As Required
    ("Coating System",          (27, 27)),
    ("Operator",                (27, 33)),
    ("Piggable",                (28, 21)),
    ("Diameter",                (28, 29)),
    ("Flange Rating",           (28, 33)),
    ("Flange Facing Finish",    (29, 21)),
    ("Face to Face Dimension",  (29, 31)),
    ("Nipples Both Side Ends",  (30, 21)),
    ("Fire Safe",               (30, 29)),
    ("Max Allowable Shear Torque", (31, 23)),
    ("Manufacturer",            (32, 21)),
    ("Model No.",               (32, 29)),

    # ── HYDRAULIC / ACTUATOR ─────────────────────────────────────────────────
    ("Hydraulic Supply Min",            (34, 23)),
    ("Hydraulic Supply Normal",         (34, 26)),
    ("Hydraulic Supply Max",            (34, 30)),
    ("Hydraulic Supply Design",         (34, 34)),
    ("Actuator Manufacturer",                    (35, 21)),
    ("Actuator Model",                  (35, 29)),
    ("Actuator Type",                   (36, 21)),
    ("Actuator Body Material",          (36, 29)),
    ("Max Torque",                      (37, 23)),
    ("Fluid Power",                     (37, 31)),
    ("Torque Break/Run/End",            (38, 23)),
    ("Time to Open/Close",              (38, 31)),  # Sec/Inch (a)
    ("Action on Hydraulic Failure",           (39, 21)),
    ("Action on Power Failure",               (39, 29)),
    ("Painting",                        (40, 21)),
    ("Mounting Position",               (40, 29)),
    ("Limit Switches Tag",              (41, 21)),
    ("Limit Switch Quantity",           (41, 34)),
    ("Limit Switch Manufacturer",                (42, 21)),
    ("Limit Switch Type/Model",         (42, 29)),
    ("Limit Switch Elec. Connection",   (43, 21)),
    ("Limit Switch Voltage",            (43, 31)),
    ("Pneumatic Air Supply Connection", (44, 21)),
    ("Actuator Weight",                 (45, 23)),
    ("Position Indicator",              (45, 29)),

    # ── LOCAL CONTROL PANEL ──────────────────────────────────────────────────
    ("Solenoid Valve Tags",             (46, 21)),
    ("Solenoid Valve Material",         (46, 34)),
    ("Solenoid Voltage",                (47, 23)),
    ("Solenoid Quantity",               (47, 29)),
    ("Solenoid Manufacturer",                    (48, 21)),
    ("Solenoid Model",                  (48, 29)),
    ("Sol. Elec. Connection",           (49, 21)),
    ("Sol. Consumption",                (49, 31)),
    ("Hydraulic Timer",                 (50, 21)),
    ("Tubing / Fitting",                (50, 29)),
    ("Reset Push Button Tag",           (51, 21)),
    ("Reset Elec. Connection",          (51, 29)),
    ("Solenoid Valve Test",             (52, 21)),
    ("Actuator Air Pressure Supply",    (53, 21)),
    ("Actuator Electrical Connection",       (53, 29)),
    ("Partial Stroking Test",           (54, 21)),
    ("Local Open/Close Control",        (55, 21)),
    ("Local Control Panel Material",                    (56, 21)),
    ("Typical (PID)",               (56, 29)),
    ("Local Control Panel Mounting",                    (57, 21)),
    ("Local Control Panel Weight",                      (58, 23)),
    ("Local Control Panel Manufacturer",                (58, 29)),

    # ── PURCHASE ─────────────────────────────────────────────────────────────
    ("Purchase Manufacturer",                    (59, 21)),
    ("Purchase Model",                  (59, 27)),
    ("Purchase Order Number",           (59, 33)),
    ("Item Number",                     (60, 21)),
    ("Serial Number",                   (60, 29)),

]


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def clean(value) -> str:
    """Convert any cell value to a clean string."""
    if value is None:
        return ""
    return str(value).strip()


def is_spec_sheet(name: str) -> bool:
    if name.strip().lower() in {s.lower() for s in SKIP_SHEETS}:
        return False
    return bool(re.match(r'^\d{2}-[A-Za-z]+-', name.strip()))


def extract_row(sheet) -> list:
    """Read all FIELDS from one xlwings sheet object."""
    row = []
    for _, (r, c) in FIELDS:
        try:
            val = sheet.cells(r, c).value
            row.append(clean(val))
        except Exception:
            row.append("")
    return row


def style_header(ws) -> None:
    fill   = PatternFill(fill_type='solid', fgColor='1F4E78')
    font   = Font(color='FFFFFF', bold=True)
    thin   = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border


def style_data(ws) -> None:
    thin   = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border    = border


def autosize(ws) -> None:
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            for line in str(cell.value or "").split('\n'):
                max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main() -> None:
    print('=' * 60)
    print('SPEC SHEET DATA EXTRACTOR — xlwings version')
    print('=' * 60)
    print()
    print('This script reads directly from Excel while it is open.')
    print('Make sure your .xlsm file is already open in Excel.')
    print()

    # List currently open workbooks via xlwings
    try:
        app = xw.apps.active
        if app is None:
            raise RuntimeError("No Excel application found.")
        open_books = [b.name for b in app.books]
    except Exception as e:
        print(f"ERROR: Could not connect to Excel. Is Excel open?")
        print(f"  Detail: {e}")
        input("\nPress Enter to exit...")
        sys.exit(1)

    # Filter out temporary/system workbooks (no sheets matching spec pattern)
    xlsm_books = [b for b in app.books if b.name.lower().endswith(('.xlsm', '.xlsx'))]

    if len(xlsm_books) == 1:
        wb_xw = xlsm_books[0]
        print(f"Auto-detected workbook: {wb_xw.name}")
        print()
    else:
        print("Multiple workbooks open in Excel:")
        for name in open_books:
            print(f"  - {name}")
        print()
        filename = input("Enter the filename to process (e.g. project_file.xlsm): ").strip()
        if not filename:
            print("ERROR: No filename entered.")
            input("\nPress Enter to exit...")
            sys.exit(1)
        wb_xw = None
        for book in app.books:
            if book.name.lower() == filename.lower():
                wb_xw = book
                break
        if wb_xw is None:
            print(f"ERROR: '{filename}' is not currently open in Excel.")
            print("Please open it in Excel first, then run this script again.")
            input("\nPress Enter to exit...")
            sys.exit(1)
        print(f"Connected to: {wb_xw.name}")
        print()

    # Filter spec sheets
    spec_sheets = [s for s in wb_xw.sheets if is_spec_sheet(s.name)]
    if not spec_sheets:
        print("ERROR: No spec sheets detected (expected names like 21-SDV-211001).")
        input("\nPress Enter to exit...")
        sys.exit(1)

    print(f"Found {len(spec_sheets)} spec sheet(s). Extracting...")

    # Build output workbook with openpyxl
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = OUTPUT_SHEET_NAME

    headers = [h for h, _ in FIELDS]
    out_ws.append(headers)

    for sheet in spec_sheets:
        out_ws.append(extract_row(sheet))
        print(f"  Extracted : {sheet.name}")

    style_header(out_ws)
    style_data(out_ws)
    out_ws.freeze_panes = 'A2'
    out_ws.auto_filter.ref = out_ws.dimensions
    autosize(out_ws)

    # Save output next to this script
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    base        = os.path.splitext(wb_xw.name)[0]
    output_file = os.path.join(script_dir, f"{base}_consolidated.xlsx")
    out_wb.save(output_file)

    print()
    print('=' * 60)
    print(f"Done!")
    print(f"Output file : {output_file}")
    print(f"Sheets      : {len(spec_sheets)}")
    print(f"Columns     : {len(headers)}")
    print('=' * 60)
    input("\nPress Enter to exit...")


if __name__ == '__main__':
    main()
